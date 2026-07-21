from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "docs" / "Idle_Balance_Bible.xlsx"

GREEN = "2E7D32"
DARK = "17351F"
GOLD = "D6A323"
LIGHT = "E8F5E9"
GRAY = "ECEFF1"
RED = "C62828"


def setup_sheet(ws, title, headers):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK)
    if headers:
        ws.insert_rows(2)
        for index, value in enumerate(headers, 1):
            cell = ws.cell(2, index, value)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=GREEN)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A3"


def autosize(ws, maximum=34):
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        width = max(len(str(cell.value or "")) for cell in column) + 2
        ws.column_dimensions[letter].width = min(max(width, 10), maximum)


def style_data(ws, start_row=3):
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            if cell.row % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=LIGHT)
            cell.alignment = Alignment(vertical="top")


wb = Workbook()
wb.remove(wb.active)

# README
ws = wb.create_sheet("README")
ws.sheet_view.showGridLines = False
rows = [
    ("Idle Balance Bible", "Design authority for tunable progression and economy."),
    ("Target", "Focused free player reaches Node Exploration Lv.100 in 28–32 days."),
    ("Level model", "Each Production Node owns an independent Exploration Level."),
    ("Lv.1–100", "Vanilla pools only."),
    ("Lv.101+", "Adds MMORPG materials while retaining vanilla output."),
    ("Yellow cells", "Primary tuning assumptions."),
    ("Formula cells", "Derived outputs; do not replace without updating the model."),
    ("Credits", "1 Credit purchase price = 1 THB; non-cashable and non-transferable."),
    ("Important", "Weights do not replace daily/weekly rare-item caps."),
]
for r, values in enumerate(rows, 1):
    for c, value in enumerate(values, 1):
        ws.cell(r, c, value)
ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
ws["A1"].fill = PatternFill("solid", fgColor=DARK)
ws["B1"].fill = PatternFill("solid", fgColor=DARK)
ws["B1"].font = Font(color="FFFFFF")
autosize(ws, 70)

# Assumptions
ws = wb.create_sheet("Assumptions")
setup_sheet(ws, "Core assumptions", ["Key", "Value", "Unit", "Design note"])
assumptions = [
    ("xp_base", 120, "EXP", "EXP_next = base + step × current level"),
    ("xp_step", 12, "EXP/level", "Stable curve"),
    ("passive_typical", 650, "EXP/day", "Focused node with normal check-ins"),
    ("passive_cap", 900, "EXP/day", "Per node"),
    ("first_collect", 100, "EXP/day", "Focused node"),
    ("focused_commission", 400, "EXP/day", "Focused node"),
    ("first_expedition", 700, "EXP/day", "Focused node"),
    ("extra_expedition_cap", 300, "EXP/day", "100 per additional completion"),
    ("weekly_chapter", 3500, "EXP/week", "Focused node"),
    ("tutorial", 2000, "EXP", "One time"),
    ("buffer_per_tier", 256, "items", "Target 2–3 check-ins/day"),
    ("full_research_factor", 0.25, "ratio", "Research while buffer full"),
    ("credit_coin_offset", 20, "Coins/Credit", "Checkout offset only"),
    ("hybrid_pay_cap", 0.15, "ratio", "Per qualifying checkout"),
]
for row in assumptions:
    ws.append(row)
for row in ws.iter_rows(min_row=3, max_col=4):
    row[1].fill = PatternFill("solid", fgColor="FFF3CD")
style_data(ws)
autosize(ws, 54)

# XP curve
ws = wb.create_sheet("XP Curve")
setup_sheet(ws, "Exploration Level curve", ["Current level", "EXP to next", "Cumulative EXP", "Milestone"])
cum = 0
for level in range(1, 101):
    exp_next = 0 if level == 100 else 120 + 12 * level
    milestone = ""
    if level in (10, 25, 50, 75, 100):
        milestone = {10: "Bracket II", 25: "Specialization I", 50: "Specialization II",
                     75: "Mastery perk", 100: "Capstone / Frontier eligible"}[level]
    ws.append((level, exp_next, cum, milestone))
    cum += exp_next
style_data(ws)
ws.conditional_formatting.add(
    "A3:A102",
    CellIsRule(operator="equal", formula=["100"], fill=PatternFill("solid", fgColor=GOLD)),
)
autosize(ws)

# 30-day simulation
ws = wb.create_sheet("30 Day Simulation")
headers = ["Day", "Passive", "Collect", "Commission", "First expedition",
           "Extra expedition", "Weekly", "Tutorial", "Daily total", "Cumulative", "Projected level"]
setup_sheet(ws, "Focused-node 30-day simulation", headers)
for day in range(1, 31):
    row = day + 2
    weekly = 3500 if day in (7, 14, 21, 28) else 0
    tutorial = 2000 if day == 1 else 0
    ws.append((day, 650, 100, 400, 700, 0, weekly, tutorial,
               f"=SUM(B{row}:H{row})",
               f"=IF(A{row}=1,I{row},J{row-1}+I{row})",
               f'=LOOKUP(J{row},\'XP Curve\'!$C$3:$C$102,\'XP Curve\'!$A$3:$A$102)'))
style_data(ws)
for col in range(2, 9):
    for row in range(3, 33):
        ws.cell(row, col).fill = PatternFill("solid", fgColor="FFF3CD")
autosize(ws)

# Production
ws = wb.create_sheet("Production")
headers = ["Node", "Tier", "Workers", "Base/hour", "Rarity power", "Worker level",
           "Avg diligence", "Estimated/hour", "Buffer", "Hours to full"]
setup_sheet(ws, "Production and buffer model", headers)
base_rates = {"Mining": 40, "Farming": 48, "Woodcutting": 44, "Livestock": 36, "Hunter": 30}
for node, base in base_rates.items():
    for tier in range(1, 6):
        row = ws.max_row + 1
        ws.append((node, tier, tier, base, 1.0, 10, 10,
                   f"=D{row}*C{row}*E{row}*(1+F{row}*0.02)*(1+G{row}/100)",
                   f"=256*B{row}", f"=I{row}/H{row}"))
style_data(ws)
for row in range(3, ws.max_row + 1):
    for col in (4, 5, 6, 7, 9):
        ws.cell(row, col).fill = PatternFill("solid", fgColor="FFF3CD")
autosize(ws)

# Coin economy
ws = wb.create_sheet("Coin Economy")
setup_sheet(ws, "Month-one Coin economy", ["Category", "Source/Sink", "Monthly Coins", "Assumption"])
coin_rows = [
    ("Income", "Online payout", 36000, "60 hours × 600/hour"),
    ("Income", "Daily behavior commissions", 18000, "600/day"),
    ("Income", "Weekly chapters", 8000, "2,000/week"),
    ("Income", "Achievements", 10000, "Journey/Mastery budget"),
    ("Income", "Expected Global Expedition", 2000, "Participation expectation"),
    ("Sink", "Production claims", -4000, "Four after free starter"),
    ("Sink", "Primary T1→T5", -11872, "Exact current formula"),
    ("Sink", "Three secondary T2", -3000, "1,000 each"),
    ("Sink", "20 hires", -5000, "250 each"),
    ("Sink", "Fuse catalysts/respec", -6000, "Expected"),
    ("Sink", "Warehouse expansion", -5000, "One purchase"),
    ("Sink", "Bag/utility/cosmetics", -8000, "Player choice"),
]
for item in coin_rows:
    ws.append(item)
total_row = ws.max_row + 1
ws.append(("Net", "Month-one remainder", f"=SUM(C3:C{total_row-1})", "Choice budget"))
style_data(ws)
ws.cell(total_row, 1).fill = PatternFill("solid", fgColor=GOLD)
ws.cell(total_row, 2).fill = PatternFill("solid", fgColor=GOLD)
ws.cell(total_row, 3).fill = PatternFill("solid", fgColor=GOLD)
autosize(ws)

# Credits
ws = wb.create_sheet("Credits")
setup_sheet(ws, "Hybrid Pay rules", ["Rule", "Value", "Purpose"])
credit_rows = [
    ("Purchase price", "1 Credit = 1 THB", "External checkout price, not redemption value"),
    ("Coin offset", "20 Coins/Credit", "Checkout only"),
    ("Per-purchase cap", "15%", "At least 85% earned Coins"),
    ("Season cap", "30,000 Coin offset", "Also ≤25% of Coins earned"),
    ("Never allowed", "EXP, RNG, tradable items, MMORPG materials", "Fairness"),
    ("Transfer/cashout", "Never", "Server entitlement only"),
]
for item in credit_rows:
    ws.append(item)
style_data(ws)
autosize(ws, 64)

drop_data = {
    "Mining Drops": [
        ("COBBLESTONE", 1, [60, 45, 35, 28, 22, 18, 15, 12, 10, 8], "", "", "Passive"),
        ("STONE", 1, [25, 20, 17, 14, 12, 10, 8, 7, 6, 5], "", "", "Passive"),
        ("COAL", 1, [15, 20, 20, 18, 16, 14, 12, 10, 9, 8], "", "", "Passive"),
        ("DEEPSLATE", 2, [0, 10, 10, 9, 8, 7, 6, 5, 4, 4], "", "", "Passive"),
        ("RAW_COPPER", 2, [0, 5, 8, 8, 7, 6, 5, 5, 4, 4], "", "", "Passive"),
        ("RAW_IRON", 3, [0, 0, 8, 12, 15, 17, 18, 18, 18, 18], "", "", "Passive"),
        ("FLINT", 3, [0, 0, 2, 2, 2, 2, 2, 2, 2, 2], "", "", "Passive"),
        ("REDSTONE", 4, [0, 0, 0, 5, 8, 10, 12, 13, 14, 14], "", "", "Passive"),
        ("RAW_GOLD", 4, [0, 0, 0, 3, 5, 7, 8, 9, 10, 10], "", "", "Passive"),
        ("LAPIS_LAZULI", 5, [0, 0, 0, 0, 5, 7, 8, 9, 10, 10], "", "", "Passive"),
        ("QUARTZ", 5, [0, 0, 0, 0, 4, 5, 6, 7, 8, 8], "", "", "Passive"),
        ("AMETHYST_SHARD", 6, [0, 0, 0, 0, 0, 3, 4, 5, 6, 7], "", "", "Passive"),
        ("DIAMOND", 6, [0, 0, 0, 0, 0, 1, 1.5, 2, 2.5, 3], "8/16", "", "Passive capped"),
        ("EMERALD", 7, [0, 0, 0, 0, 0, 0, 1, 1.5, 2, 2.5], "12", "", "Passive capped"),
        ("OBSIDIAN", 8, [0, 0, 0, 0, 0, 0, 0, 1, 1.5, 2], "", "", "Passive"),
        ("ANCIENT_DEBRIS", 10, [0, 0, 0, 0, 0, 0, 0, 0, 0, 0.2], "", "2 shared", "Passive capped"),
    ],
    "Farming Drops": [
        ("WHEAT", 1, [50, 40, 32, 27, 23, 20, 18, 16, 14, 12], "", "", "Passive"),
        ("CARROT", 1, [30, 25, 22, 19, 17, 15, 14, 13, 12, 11], "", "", "Passive"),
        ("POTATO", 1, [20, 20, 18, 16, 14, 13, 12, 11, 10, 9], "", "", "Passive"),
        ("BEETROOT", 2, [0, 8, 9, 9, 8, 8, 7, 7, 6, 6], "", "", "Passive"),
        ("PUMPKIN", 2, [0, 4, 6, 7, 7, 7, 7, 7, 7, 7], "", "", "Passive"),
        ("MELON_SLICE", 2, [0, 3, 5, 6, 6, 6, 6, 6, 6, 6], "", "", "Passive"),
        ("SUGAR_CANE", 3, [0, 0, 6, 8, 9, 10, 10, 10, 10, 10], "", "", "Passive"),
        ("CACTUS", 3, [0, 0, 3, 4, 5, 5, 5, 5, 5, 5], "", "", "Passive"),
        ("COCOA_BEANS", 3, [0, 0, 3, 4, 5, 5, 5, 5, 5, 5], "", "", "Passive"),
        ("BROWN_MUSHROOM", 4, [0, 0, 0, 4, 5, 5, 6, 6, 6, 6], "", "", "Passive"),
        ("SWEET_BERRIES", 4, [0, 0, 0, 4, 5, 6, 6, 6, 6, 6], "", "", "Passive"),
        ("BAMBOO", 4, [0, 0, 0, 4, 5, 6, 6, 6, 6, 6], "", "", "Passive"),
        ("NETHER_WART", 5, [0, 0, 0, 0, 0, 1, 1.5, 2, 2.5, 3], "", "", "Event until B6"),
        ("HONEY_BOTTLE", 6, [0, 0, 0, 0, 0, 2, 3, 4, 5, 6], "", "", "Passive"),
        ("CHORUS_FRUIT", 8, [0, 0, 0, 0, 0, 0, 0, 1, 2, 3], "", "", "Event/Passive"),
    ],
    "Wood Drops": [
        ("OAK_LOG", 1, [45, 35, 28, 24, 21, 18, 16, 14, 12, 10], "", "", "Passive"),
        ("BIRCH_LOG", 1, [30, 25, 21, 18, 16, 14, 13, 12, 11, 10], "", "", "Passive"),
        ("SPRUCE_LOG", 1, [25, 23, 20, 18, 16, 15, 14, 13, 12, 11], "", "", "Passive"),
        ("JUNGLE_LOG", 2, [0, 9, 11, 12, 12, 12, 12, 12, 12, 12], "", "", "Passive"),
        ("ACACIA_LOG", 2, [0, 8, 10, 11, 11, 11, 11, 11, 11, 11], "", "", "Passive"),
        ("DARK_OAK_LOG", 3, [0, 0, 10, 12, 13, 14, 14, 14, 14, 14], "", "", "Passive"),
        ("MANGROVE_LOG", 3, [0, 0, 6, 8, 9, 10, 10, 10, 10, 10], "", "", "Passive"),
        ("CHERRY_LOG", 4, [0, 0, 0, 8, 10, 11, 12, 12, 12, 12], "", "", "Passive"),
        ("BAMBOO_BLOCK", 4, [0, 0, 0, 5, 7, 8, 9, 9, 9, 9], "", "", "Passive"),
        ("CRIMSON_STEM", 6, [0, 0, 0, 0, 0, 2, 3, 4, 5, 6], "", "", "Event/Passive"),
        ("WARPED_STEM", 6, [0, 0, 0, 0, 0, 2, 3, 4, 5, 6], "", "", "Event/Passive"),
    ],
    "Livestock Drops": [
        ("BEEF", 1, [30, 25, 21, 18, 16, 14, 13, 12, 11, 10], "", "", "Passive"),
        ("PORKCHOP", 1, [25, 22, 19, 17, 15, 14, 13, 12, 11, 10], "", "", "Passive"),
        ("CHICKEN", 1, [20, 18, 16, 14, 13, 12, 11, 10, 9, 8], "", "", "Passive"),
        ("EGG", 1, [25, 20, 17, 15, 13, 12, 11, 10, 9, 8], "", "", "Passive"),
        ("LEATHER", 2, [0, 9, 11, 12, 13, 14, 15, 15, 15, 15], "", "", "Passive"),
        ("WHITE_WOOL", 2, [0, 7, 9, 10, 11, 12, 13, 13, 13, 13], "", "", "Passive"),
        ("FEATHER", 2, [0, 5, 6, 7, 7, 7, 7, 7, 7, 7], "", "", "Passive"),
        ("MILK_BUCKET", 3, [0, 0, 2, 3, 4, 5, 5, 5, 5, 5], "", "", "Passive"),
        ("RABBIT_HIDE", 3, [0, 0, 3, 4, 5, 6, 6, 6, 6, 6], "", "", "Passive"),
        ("COD", 4, [0, 0, 0, 7, 8, 9, 9, 9, 9, 9], "", "", "Passive"),
        ("SALMON", 4, [0, 0, 0, 5, 6, 7, 7, 7, 7, 7], "", "", "Passive"),
        ("GLOW_INK_SAC", 5, [0, 0, 0, 0, 2, 3, 4, 5, 6, 7], "", "", "Passive"),
        ("RABBIT_FOOT", 6, [0, 0, 0, 0, 0, 1, 1.5, 2, 2.5, 3], "", "", "Passive"),
        ("NAUTILUS_SHELL", 6, [0, 0, 0, 0, 0, 0.5, 0.7, 1, 1.2, 1.5], "4", "", "Capped"),
    ],
    "Hunter Drops": [
        ("ROTTEN_FLESH", 1, [40, 32, 26, 21, 18, 15, 13, 11, 10, 9], "", "", "Passive"),
        ("BONE", 1, [35, 29, 25, 22, 20, 18, 17, 16, 15, 14], "", "", "Passive"),
        ("STRING", 1, [25, 22, 19, 17, 15, 14, 13, 12, 11, 10], "", "", "Passive"),
        ("GUNPOWDER", 2, [0, 9, 12, 14, 15, 16, 17, 18, 18, 18], "", "", "Passive"),
        ("SPIDER_EYE", 2, [0, 6, 8, 9, 10, 10, 10, 10, 10, 10], "", "", "Passive"),
        ("ARROW", 2, [0, 5, 6, 7, 7, 7, 7, 7, 7, 7], "", "", "Passive"),
        ("SLIME_BALL", 3, [0, 0, 5, 7, 8, 9, 10, 10, 10, 10], "", "", "Passive"),
        ("MAGMA_CREAM", 3, [0, 0, 2, 3, 4, 5, 6, 7, 8, 8], "", "", "Passive"),
        ("ENDER_PEARL", 4, [0, 0, 0, 1, 1.5, 2, 2.5, 3, 3.5, 4], "", "", "Passive"),
        ("BLAZE_ROD", 5, [0, 0, 0, 0, 0, 1, 1.5, 2, 2.5, 3], "", "", "Event until B6"),
        ("GHAST_TEAR", 6, [0, 0, 0, 0, 0, 0.5, 0.7, 1, 1.2, 1.5], "4", "", "Capped"),
        ("PHANTOM_MEMBRANE", 6, [0, 0, 0, 0, 0, 1, 1.5, 2, 2.5, 3], "", "", "Passive"),
        ("PRISMARINE_SHARD", 8, [0, 0, 0, 0, 0, 0, 0, 1, 1.5, 2], "", "", "Event/Passive"),
        ("WITHER_SKELETON_SKULL", 10, [0] * 10, "", "1", "Active event only"),
    ],
}

drop_headers = ["Material", "Unlock bracket"] + [f"B{i} weight" for i in range(1, 11)] + ["Daily cap", "Weekly cap", "Source"]
for sheet_name, items in drop_data.items():
    ws = wb.create_sheet(sheet_name)
    setup_sheet(ws, sheet_name, drop_headers)
    for material, unlock, weights, daily, weekly, source in items:
        ws.append((material, unlock, *weights, daily, weekly, source))
    style_data(ws)
    for row in range(3, ws.max_row + 1):
        ws.cell(row, 2).fill = PatternFill("solid", fgColor="FFF3CD")
        for col in range(3, 13):
            if ws.cell(row, col).value:
                ws.cell(row, col).fill = PatternFill("solid", fgColor="FFF3CD")
    autosize(ws)

# Events
ws = wb.create_sheet("Event Families")
setup_sheet(ws, "Exploration event families", ["Family", "Primary stat", "Base duration", "Node EXP", "Main reward", "First bracket"])
events = [
    ("Survey", "Stamina", 15, 700, "Node EXP", 2),
    ("Resource Run", "Diligence", 25, 250, "Quantity", 2),
    ("Rare Discovery", "Luck", 30, 300, "Rare pool", 4),
    ("Rescue", "Speed", 12, 350, "Chronicle/social", 4),
    ("Endurance", "Stamina + team", 45, 400, "Worker EXP", 6),
    ("Deep Expedition", "Balanced", 60, 600, "Capstone/mastery", 8),
]
for event in events:
    ws.append(event)
style_data(ws)
autosize(ws)

# Post-100
ws = wb.create_sheet("Post 100")
setup_sheet(ws, "Post-100 roll shares", ["Level range", "Vanilla share", "MMORPG share", "Material tier"])
post_rows = [
    ("101–110", 0.90, 0.10, "Frontier I"),
    ("111–125", 0.80, 0.20, "Frontier I"),
    ("126–150", 0.70, 0.30, "Frontier I–II"),
    ("151–175", 0.60, 0.40, "Frontier II"),
    ("176–200", 0.50, 0.50, "Frontier II–III"),
]
for item in post_rows:
    ws.append(item)
for row in range(3, ws.max_row + 1):
    ws.cell(row, 2).number_format = "0%"
    ws.cell(row, 3).number_format = "0%"
style_data(ws)
autosize(ws)

# Achievement budget
ws = wb.create_sheet("Achievement Budget")
setup_sheet(ws, "Month-one Achievement budget", ["Reward", "Minimum", "Maximum", "Rule"])
achievement_rows = [
    ("Exploration EXP", 1500, 2000, "Journey only, non-repeatable"),
    ("Coins", 8000, 10000, "Journey and Mastery"),
    ("Chronicle Points", 40, 60, "Permanent account score"),
    ("Cosmetics", 6, 10, "Titles, façades, frames"),
    ("Functional presets", 1, 2, "Organization convenience"),
]
for item in achievement_rows:
    ws.append(item)
style_data(ws)
autosize(ws)

# Node-type perks
ws = wb.create_sheet("Node Type Perks")
setup_sheet(ws, "Node-specific perk milestones", ["Node", "Unlock level", "Tier", "Perk", "Primary function"])
perk_rows = {
    "Mining": [
        (15, "Foundation", "Stone Mason", "Common-block preference/project routing"),
        (15, "Foundation", "Ore Sense", "30-roll basic-ore pity"),
        (15, "Foundation", "Reinforced Shaft", "Buffer/full-research"),
        (35, "Operations", "Seismic Map", "Daily event choice"),
        (35, "Operations", "Rich Vein", "Prepared quantity"),
        (35, "Operations", "Crystal Echo", "Capped-roll fallback"),
        (60, "Direction", "Metal Stratum", "Metal preference"),
        (60, "Direction", "Redstone Stratum", "Technical-resource preference"),
        (60, "Direction", "Gem Stratum", "Gem/event preference"),
        (85, "Legacy", "Deep Core", "Advanced-roll meter"),
        (85, "Legacy", "Excavation Crew", "Reduced event opportunity cost"),
        (85, "Legacy", "Geological Archive", "Journal research"),
    ],
    "Farming": [
        (15, "Foundation", "Crop Rotation", "Variety research"),
        (15, "Foundation", "Seed Keeper", "Common-crop bonus"),
        (15, "Foundation", "Irrigation Channels", "Buffer/daily collection"),
        (35, "Operations", "Greenhouse", "Early full-buffer research"),
        (35, "Operations", "Market Garden", "Commission efficiency"),
        (35, "Operations", "Pollinator Route", "Prepared advanced roll"),
        (60, "Direction", "Staple Harvest", "Staple preference"),
        (60, "Direction", "Exotic Garden", "Exotic preference"),
        (60, "Direction", "Alchemist's Plot", "Potion-resource preference"),
        (85, "Legacy", "Perennial Field", "Daily regrowth"),
        (85, "Legacy", "Harvest Festival", "Weekly collection event"),
        (85, "Legacy", "Living Soil", "Rested research"),
    ],
    "Woodcutting": [
        (15, "Foundation", "Sustainable Grove", "Species cycle"),
        (15, "Foundation", "Lumber Stacks", "Log buffer"),
        (15, "Foundation", "Sapling Keeper", "Bound cosmetic tokens"),
        (35, "Operations", "Trailblazer", "Event speed"),
        (35, "Operations", "Carpenter's Measure", "Project efficiency"),
        (35, "Operations", "Forest Memory", "Journal Node EXP"),
        (60, "Direction", "Temperate Grove", "Temperate preference"),
        (60, "Direction", "Wild Grove", "Wild-biome preference"),
        (60, "Direction", "Otherworld Grove", "Nether-wood preference"),
        (85, "Legacy", "Elder Grove", "Daily route choice"),
        (85, "Legacy", "Master Carpenter", "Construction bundles"),
        (85, "Legacy", "Heartwood", "Advanced-roll pity"),
    ],
    "Livestock": [
        (15, "Foundation", "Balanced Herd", "Diversity Node EXP"),
        (15, "Foundation", "Feed Reserve", "Buffer/preparation"),
        (15, "Foundation", "Ranch Hand", "Worker EXP"),
        (35, "Operations", "Breeding Records", "Daily targeting"),
        (35, "Operations", "Fishery Route", "Aquatic events"),
        (35, "Operations", "Shepherd's Call", "Role synergy"),
        (60, "Direction", "Ranch Table", "Food preference"),
        (60, "Direction", "Weaver's Pasture", "Textile preference"),
        (60, "Direction", "River Keeper", "Aquatic preference"),
        (85, "Legacy", "Legendary Herd", "Weekly journal event"),
        (85, "Legacy", "Sanctuary", "Remaining-crew power"),
        (85, "Legacy", "Ranch Legacy", "Node-bound veteran"),
    ],
    "Hunter": [
        (15, "Foundation", "Bounty Board", "Daily target family"),
        (15, "Foundation", "Scavenger", "Extra uncapped roll"),
        (15, "Foundation", "Night Watch", "Online-night events"),
        (35, "Operations", "Trail Marks", "Event speed"),
        (35, "Operations", "Trophy Sense", "Event information"),
        (35, "Operations", "Prepared Ambush", "Prepared quantity"),
        (60, "Direction", "Grave Warden", "Undead preference"),
        (60, "Direction", "Nest Breaker", "Creature preference"),
        (60, "Direction", "Rift Stalker", "Dimension preference"),
        (85, "Legacy", "Apex Contract", "Weekly route choice"),
        (85, "Legacy", "Clean Hunt", "Supply bundles"),
        (85, "Legacy", "Dread Mark", "Event completion EXP"),
    ],
}
for node, perks in perk_rows.items():
    for unlock, tier, name, function in perks:
        ws.append((node, unlock, tier, name, function))
style_data(ws)
autosize(ws, 48)

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)
print(OUTPUT)
